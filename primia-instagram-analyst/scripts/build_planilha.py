#!/usr/bin/env python3
"""
Gera planilha .xlsx final para a skill primia-instagram-analyst.

Uso:
  python build_planilha.py --modo PESADO --input dados.json --output planilha.xlsx
  python build_planilha.py --modo LEVE --input dados.json --output planilha.xlsx

Formato esperado de dados.json:
  {
    "modo": "PESADO" | "LEVE",
    "perfil_analisado": "@username",
    "n_posts_analisados": 30,
    "data_analise": "2026-05-06",
    "comentaristas": [
      {
        "username": "...",
        "link": "https://instagram.com/...",
        "profissao": "...",                # PESADO
        "confianca": "Alto|Médio|Baixo",   # PESADO
        "seguidores": 1234,                # PESADO
        "frequencia": 4,
        "qualidade_interacoes": "Alta|Média|Baixa",  # PESADO
        "eh_icp": "Sim|Parcial|Não",
        "nota": 8,                         # PESADO (0-10)
        "categoria": "Fã Recorrente ICP|Diamante Oculto|Engajador Frequente|Observar|Descartar",  # PESADO
        "temperatura": "Quente|Morno|Frio",  # LEVE
        "comentario_resumo": "...",        # LEVE
        "dm_sugerida": "...",              # LEVE
        "observacoes": "..."
      }
    ]
  }
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Erro: openpyxl não instalado. Rode: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)


# ============ ESTILOS ============
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Cores por categoria (PESADO)
CATEGORIA_FILLS = {
    "Fã Recorrente ICP": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),  # verde
    "Diamante Oculto":   PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # azul claro
    "Engajador Frequente": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),  # amarelo
    "Observar":          PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # cinza
    "Descartar":         PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),  # laranja claro
}

# Cores por temperatura (LEVE)
TEMPERATURA_FILLS = {
    "Quente": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),  # laranja-vermelho claro
    "Morno":  PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),  # amarelo
    "Frio":   PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),  # azul gelo
}

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Ordem de categorias (para sort)
ORDEM_CATEGORIA = [
    "Fã Recorrente ICP",
    "Diamante Oculto",
    "Engajador Frequente",
    "Observar",
    "Descartar",
]
ORDEM_TEMPERATURA = ["Quente", "Morno", "Frio"]


def build_pesado(wb, dados):
    """Planilha completa, 11 colunas."""
    ws = wb.active
    ws.title = "Comentaristas"

    headers = [
        "Username",
        "Link do perfil",
        "Indício de profissão",
        "Confiança",
        "Seguidores",
        "Frequência (de N posts)",
        "Qualidade das interações",
        "É ICP?",
        "Nota geral (0-10)",
        "Categoria estratégica",
        "Observações",
    ]

    # Cabeçalho
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Sort: por categoria (ordem fixa), depois nota desc
    def sort_key(c):
        cat_idx = ORDEM_CATEGORIA.index(c.get("categoria", "Descartar")) if c.get("categoria") in ORDEM_CATEGORIA else 99
        nota = c.get("nota", 0)
        return (cat_idx, -nota)

    comentaristas = sorted(dados["comentaristas"], key=sort_key)

    for row_idx, c in enumerate(comentaristas, start=2):
        row_data = [
            c.get("username", ""),
            c.get("link", ""),
            c.get("profissao", ""),
            c.get("confianca", ""),
            c.get("seguidores", ""),
            c.get("frequencia", ""),
            c.get("qualidade_interacoes", ""),
            c.get("eh_icp", ""),
            c.get("nota", ""),
            c.get("categoria", ""),
            c.get("observacoes", ""),
        ]

        categoria = c.get("categoria", "")
        fill = CATEGORIA_FILLS.get(categoria)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill and col_idx == 10:  # destaca só a coluna Categoria
                cell.fill = fill

        # Hyperlink na coluna Link
        link = c.get("link", "")
        if link:
            ws.cell(row=row_idx, column=2).hyperlink = link
            ws.cell(row=row_idx, column=2).font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    # Larguras
    larguras = [18, 30, 24, 11, 12, 14, 16, 10, 12, 22, 40]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze cabeçalho
    ws.freeze_panes = "A2"

    # Filtro automático
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Aba de resumo
    add_resumo_pesado(wb, dados, comentaristas)


def add_resumo_pesado(wb, dados, comentaristas):
    """Aba de resumo executivo."""
    ws = wb.create_sheet(title="Resumo", index=0)

    ws["A1"] = "Mapeamento de Comentaristas — Resumo Executivo"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    ws.merge_cells("A1:C1")

    ws["A3"] = "Perfil analisado:"
    ws["B3"] = dados.get("perfil_analisado", "")
    ws["A4"] = "Posts analisados:"
    ws["B4"] = dados.get("n_posts_analisados", "")
    ws["A5"] = "Data da análise:"
    ws["B5"] = dados.get("data_analise", "")
    ws["A6"] = "Total de comentaristas únicos:"
    ws["B6"] = len(comentaristas)

    for row in range(3, 7):
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)

    # Contagem por categoria
    ws["A8"] = "Categoria estratégica"
    ws["B8"] = "Quantidade"
    ws["C8"] = "% do total"
    for col in ["A8", "B8", "C8"]:
        ws[col].fill = HEADER_FILL
        ws[col].font = HEADER_FONT
        ws[col].alignment = HEADER_ALIGN

    total = len(comentaristas) or 1
    for i, cat in enumerate(ORDEM_CATEGORIA, start=9):
        qtd = sum(1 for c in comentaristas if c.get("categoria") == cat)
        ws[f"A{i}"] = cat
        ws[f"B{i}"] = qtd
        ws[f"C{i}"] = f"{qtd/total*100:.1f}%"
        fill = CATEGORIA_FILLS.get(cat)
        if fill:
            ws[f"A{i}"].fill = fill
        for col in ["A", "B", "C"]:
            ws[f"{col}{i}"].border = THIN_BORDER

    # Próximos passos
    ws["A16"] = "Próximos passos sugeridos"
    ws["A16"].font = Font(name="Calibri", size=12, bold=True, color="1F4E79")

    fas = sum(1 for c in comentaristas if c.get("categoria") == "Fã Recorrente ICP")
    diamantes = sum(1 for c in comentaristas if c.get("categoria") == "Diamante Oculto")

    txt_fas = f"o {fas} perfil em 'Fã Recorrente ICP'" if fas == 1 else f"os {fas} perfis em 'Fã Recorrente ICP'"
    txt_dia = f"o {diamantes} 'Diamante Oculto'" if diamantes == 1 else f"os {diamantes} 'Diamantes Ocultos'"

    ws["A17"] = f"1. Abordar com DM personalizada {txt_fas} nesta semana."
    ws["A18"] = f"2. Engajar nos posts d{'o' if diamantes == 1 else 'os'} {txt_dia} por 2-3 semanas antes de abordar."
    ws["A19"] = "3. Não abordar 'Engajador Frequente' — manter como audiência fiel."
    ws["A20"] = "4. Reavaliar 'Observar' na próxima rodada mensal."

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14


def build_leve(wb, dados):
    """Planilha enxuta, 6 colunas."""
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "Username",
        "Link do perfil",
        "É ICP?",
        "Temperatura",
        "O que ele comentou",
        "Sugestão de primeira DM",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    def sort_key(c):
        temp = c.get("temperatura", "Frio")
        idx = ORDEM_TEMPERATURA.index(temp) if temp in ORDEM_TEMPERATURA else 99
        return idx

    comentaristas = sorted(dados["comentaristas"], key=sort_key)

    for row_idx, c in enumerate(comentaristas, start=2):
        row_data = [
            c.get("username", ""),
            c.get("link", ""),
            c.get("eh_icp", ""),
            c.get("temperatura", ""),
            c.get("comentario_resumo", ""),
            c.get("dm_sugerida", ""),
        ]

        temp = c.get("temperatura", "")
        fill = TEMPERATURA_FILLS.get(temp)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill and col_idx == 4:
                cell.fill = fill

        link = c.get("link", "")
        if link:
            ws.cell(row=row_idx, column=2).hyperlink = link
            ws.cell(row=row_idx, column=2).font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    larguras = [18, 30, 12, 14, 40, 50]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def main():
    parser = argparse.ArgumentParser(description="Gera planilha .xlsx para primia-instagram-analyst")
    parser.add_argument("--modo", choices=["PESADO", "LEVE"], required=True)
    parser.add_argument("--input", required=True, help="Arquivo JSON com os dados")
    parser.add_argument("--output", required=True, help="Caminho do .xlsx de saída")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        dados = json.load(f)

    wb = Workbook()

    if args.modo == "PESADO":
        build_pesado(wb, dados)
    else:
        build_leve(wb, dados)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"✓ Planilha gerada: {args.output}")


if __name__ == "__main__":
    main()
