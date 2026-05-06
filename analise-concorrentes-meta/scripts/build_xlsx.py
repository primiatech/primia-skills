#!/usr/bin/env python3
"""
build_xlsx.py — Constrói a planilha de análise competitiva.

Uso:
    python build_xlsx.py \
        --enriched /home/claude/output/concorrente1/enriched_ads.json \
        /home/claude/output/concorrente2/enriched_ads.json \
        --analysis /home/claude/output/analysis.json \
        --output /mnt/user-data/outputs/analise_concorrentes.xlsx

`analysis.json` é gerado por VOCÊ (Claude) na Etapa 5 e contém:
{
    "by_competitor": { "competitor_slug": { "summary": "...", "top_angles": [...], ... } },
    "comparative_matrix": [...],
    "recommendations": [...]
}

A planilha terá: Resumo, Anúncios, Top Longevos, Comparativo, Recomendações.

IMPORTANTE: este script só monta o arquivo. As classificações analíticas
(angulo, funil, framework, oferta) devem vir do `analysis.json` produzido pelo Claude,
ou de campos `analysis` injetados em cada anúncio do enriched_ads.json.
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[col_letter]:
            try:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, min(len(v), max_width))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2


def build_summary_sheet(ws, all_data: dict):
    """Aba Resumo: estatísticas por concorrente."""
    headers = ["Concorrente", "Total de anúncios", "Ativos", "Mediana dias no ar",
               "Formato dominante", "Ângulo dominante", "Estágio dominante"]
    ws.append(headers)
    style_header(ws)

    for competitor, info in all_data.items():
        ads = info["ads"]
        analysis = info.get("analysis", {})
        active = sum(1 for a in ads if a.get("is_active"))
        days_list = [a["days_running"] for a in ads if a.get("days_running") is not None]
        median_days = (sorted(days_list)[len(days_list) // 2] if days_list else None)
        format_count = {}
        for a in ads:
            ct = a.get("creative_type", "image")
            format_count[ct] = format_count.get(ct, 0) + 1
        dom_format = max(format_count, key=format_count.get) if format_count else "—"

        ws.append([
            competitor,
            len(ads),
            active,
            median_days if median_days is not None else "—",
            dom_format,
            analysis.get("angulo_dominante", "—"),
            analysis.get("estagio_dominante", "—"),
        ])

    ws.freeze_panes = "A2"
    auto_width(ws)


def build_ads_sheet(ws, all_data: dict):
    """Aba Anúncios: linha por anúncio."""
    headers = [
        "Concorrente", "Ad ID", "Ativo", "Início", "Fim", "Dias no ar",
        "Plataformas", "Formato", "Variações", "CTA", "Link",
        "Copy", "Transcrição", "OCR",
        "Ângulo", "Funil", "Framework", "Oferta", "Notas",
    ]
    ws.append(headers)
    style_header(ws)

    for competitor, info in all_data.items():
        for ad in info["ads"]:
            ad_analysis = ad.get("analysis", {})
            ws.append([
                competitor,
                ad.get("ad_archive_id", ""),
                "Sim" if ad.get("is_active") else "Não",
                ad.get("start_date", ""),
                ad.get("end_date", ""),
                ad.get("days_running", ""),
                ", ".join(ad.get("platforms", [])),
                ad.get("creative_type", ""),
                ad.get("variations_count", 1),
                ad.get("cta_type", ""),
                ad.get("link_url", ""),
                (ad.get("body_text") or "")[:500],
                (ad.get("transcript") or "")[:500],
                (ad.get("ocr_text") or "")[:300],
                ad_analysis.get("angulo", ""),
                ad_analysis.get("funil", ""),
                ad_analysis.get("framework", ""),
                ad_analysis.get("oferta", ""),
                ad_analysis.get("notas", ""),
            ])

    ws.freeze_panes = "A2"
    auto_width(ws, max_width=50)


def build_top_longevos_sheet(ws, all_data: dict):
    """Top 20 anúncios por dias no ar (global)."""
    headers = ["Rank", "Concorrente", "Ad ID", "Dias no ar", "Formato", "CTA", "Copy (início)"]
    ws.append(headers)
    style_header(ws)

    all_ads = []
    for competitor, info in all_data.items():
        for ad in info["ads"]:
            if ad.get("days_running") is not None:
                all_ads.append((competitor, ad))

    all_ads.sort(key=lambda x: x[1]["days_running"] or 0, reverse=True)

    for i, (competitor, ad) in enumerate(all_ads[:20], 1):
        ws.append([
            i, competitor,
            ad.get("ad_archive_id", ""),
            ad.get("days_running", ""),
            ad.get("creative_type", ""),
            ad.get("cta_type", ""),
            (ad.get("body_text") or "")[:200],
        ])

    ws.freeze_panes = "A2"
    auto_width(ws)


def build_comparative_sheet(ws, analysis: dict):
    """Matriz comparativa entre concorrentes."""
    matrix = analysis.get("comparative_matrix", [])
    if not matrix:
        ws.append(["Sem dados de matriz comparativa fornecidos."])
        return

    headers = list(matrix[0].keys())
    ws.append(headers)
    style_header(ws)

    for row in matrix:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    auto_width(ws)


def build_recommendations_sheet(ws, analysis: dict):
    """Aba de recomendações."""
    headers = ["#", "Recomendação", "Evidência", "Prioridade"]
    ws.append(headers)
    style_header(ws)

    recs = analysis.get("recommendations", [])
    for i, r in enumerate(recs, 1):
        ws.append([
            i,
            r.get("titulo", ""),
            r.get("evidencia", ""),
            r.get("prioridade", "Média"),
        ])

    ws.freeze_panes = "A2"
    auto_width(ws, max_width=80)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--enriched", nargs="+", required=True,
                        help="Caminhos dos enriched_ads.json (um por concorrente)")
    parser.add_argument("--analysis", required=True,
                        help="Caminho do analysis.json gerado pelo Claude")
    parser.add_argument("--output", required=True, help="Caminho da planilha .xlsx de saída")
    args = parser.parse_args()

    # Carrega dados
    all_data = {}
    for path in args.enriched:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        competitor = d["metadata"]["advertiser"]
        all_data[competitor] = {
            "ads": d.get("ads", []),
            "analysis": {},  # preenchido abaixo se houver
        }

    with open(args.analysis, encoding="utf-8") as f:
        analysis = json.load(f)

    # Mescla análise por concorrente nos dados
    for competitor, comp_analysis in analysis.get("by_competitor", {}).items():
        if competitor in all_data:
            all_data[competitor]["analysis"] = comp_analysis

    # Monta workbook
    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb.create_sheet("Resumo"), all_data)
    build_ads_sheet(wb.create_sheet("Anúncios"), all_data)
    build_top_longevos_sheet(wb.create_sheet("Top Longevos"), all_data)
    build_comparative_sheet(wb.create_sheet("Comparativo"), analysis)
    build_recommendations_sheet(wb.create_sheet("Recomendações"), analysis)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"[✓] Planilha salva em {output_path}")


if __name__ == "__main__":
    main()
