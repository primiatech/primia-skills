#!/usr/bin/env python3
"""
build_xlsx.py

Gera a planilha base da análise. Abas:
- Resumo (1 linha por canal)
- Videos (1 linha por vídeo)
- Top_Performers (subset com hipóteses de "por que funcionou")
- Comentarios_Categorizados (1 linha por comentário com categoria)
- Voz_Avatar (consolidação por categoria)
- Comparativo (matriz competitiva)
- Recomendacoes (5-10 com evidência)

IMPORTANTE: Antes de rodar, leia /mnt/skills/public/xlsx/SKILL.md
para garantir formatação consistente.

A análise (categorização de comentários, hipóteses, recomendações) precisa ser
gerada PELO CLAUDE e passada via JSON em --analysis-json.

Uso:
    python build_xlsx.py --output-dir output/ --analysis-json analysis.json --output report.xlsx
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, max_width=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            try:
                v = str(cell.value) if cell.value else ""
                lengths.append(min(len(v), max_width))
            except Exception:
                lengths.append(0)
        if lengths:
            ws.column_dimensions[col_letter].width = min(max(lengths) + 2, max_width)


def write_resumo(ws, channels, videos_by_channel, analysis):
    ws.title = "Resumo"
    headers = [
        "Canal", "Channel ID", "Inscritos", "Vídeos analisados",
        "View ratio mediano", "Engajamento mediano",
        "Ângulo dominante", "Estágio funil dominante",
        "Formato dominante", "Postagens/sem (média)",
    ]
    ws.append(headers)

    summaries = analysis.get("channel_summaries", {})

    for ch in channels:
        slug = ch["slug"]
        videos = videos_by_channel.get(slug, [])
        s = summaries.get(slug, {})

        ratios = sorted(v["view_to_subscriber_ratio"] for v in videos)
        engagements = sorted(v["engagement_rate"] for v in videos)
        median_ratio = ratios[len(ratios) // 2] if ratios else 0
        median_eng = engagements[len(engagements) // 2] if engagements else 0

        long_count = sum(1 for v in videos if not v["is_short"])
        short_count = sum(1 for v in videos if v["is_short"])
        format_dom = "Long-form" if long_count >= short_count else "Shorts"
        if long_count == short_count and long_count > 0:
            format_dom = "Misto"

        ws.append([
            ch["title"],
            ch["channel_id"],
            ch.get("subscriber_count", 0),
            len(videos),
            round(median_ratio, 4),
            round(median_eng, 4),
            s.get("angulo_dominante", "—"),
            s.get("estagio_funil", "—"),
            format_dom,
            s.get("postagens_semana_media", "—"),
        ])

    style_header(ws)
    autosize(ws)


def write_videos(ws, videos):
    ws.title = "Videos"
    headers = [
        "Canal", "Video ID", "Título", "Publicado em", "Dias desde publicação",
        "Duração (s)", "É Short?", "Views", "Likes", "Comentários",
        "Views/dia", "View ratio", "Engajamento",
        "É Top Performer?", "Score composto",
        "Ângulo classificado", "Estágio funil", "Hook tipo",
        "URL",
    ]
    ws.append(headers)

    for v in videos:
        analysis_v = v.get("_analysis", {})
        ws.append([
            v.get("channel_title", ""),
            v["video_id"],
            v["title"],
            v["published_at"],
            v["days_since_published"],
            v["duration_seconds"],
            "Sim" if v["is_short"] else "Não",
            v["view_count"],
            v["like_count"],
            v["comment_count"],
            v["views_per_day"],
            v["view_to_subscriber_ratio"],
            v["engagement_rate"],
            "Sim" if v.get("is_top_performer") else "Não",
            v.get("composite_score", ""),
            analysis_v.get("angulo", ""),
            analysis_v.get("estagio_funil", ""),
            analysis_v.get("hook_tipo", ""),
            f"https://www.youtube.com/watch?v={v['video_id']}",
        ])

    style_header(ws)
    autosize(ws, max_width=50)


def write_top_performers(ws, top_videos_flat, analysis):
    ws.title = "Top_Performers"
    headers = [
        "Canal", "Video ID", "Título", "Categoria", "Views", "View ratio",
        "Score", "Hook (primeiros 30s)", "Promessa central", "CTA principal",
        "Por que funcionou (hipótese)",
    ]
    ws.append(headers)

    hypotheses = analysis.get("top_video_analysis", {})

    for v in top_videos_flat:
        h = hypotheses.get(v["video_id"], {})
        ws.append([
            v.get("channel_title", ""),
            v["video_id"],
            v["title"],
            v.get("format_category", ""),
            v["view_count"],
            v["view_to_subscriber_ratio"],
            v.get("composite_score", ""),
            h.get("hook_decoded", ""),
            h.get("promessa", ""),
            h.get("cta", ""),
            h.get("por_que_funcionou", ""),
        ])

    style_header(ws)
    autosize(ws, max_width=70)


def write_comentarios(ws, categorized_comments):
    ws.title = "Comentarios_Categorizados"
    headers = [
        "Canal", "Video ID", "Comment ID", "Autor", "Likes",
        "Categoria", "Texto do comentário", "Quote/Trecho extraído",
    ]
    ws.append(headers)

    for c in categorized_comments:
        ws.append([
            c.get("channel_title", ""),
            c.get("video_id", ""),
            c.get("comment_id", ""),
            c.get("author", ""),
            c.get("like_count", 0),
            c.get("categoria", ""),
            c.get("text", "")[:500],
            c.get("quote_extraido", ""),
        ])

    style_header(ws)
    autosize(ws, max_width=80)


def write_voz_avatar(ws, voz):
    ws.title = "Voz_Avatar"
    headers = ["Categoria", "Frase recorrente / Padrão", "Frequência (n menções)", "Quote literal exemplar"]
    ws.append(headers)

    for entry in voz:
        ws.append([
            entry.get("categoria", ""),
            entry.get("padrao", ""),
            entry.get("frequencia", 0),
            entry.get("quote", ""),
        ])

    style_header(ws)
    autosize(ws, max_width=80)


def write_comparativo(ws, matrix):
    ws.title = "Comparativo"
    if not matrix:
        ws.append(["—"])
        return
    headers = list(matrix[0].keys())
    ws.append(headers)
    for row in matrix:
        ws.append([row.get(h, "") for h in headers])
    style_header(ws)
    autosize(ws, max_width=50)


def write_recomendacoes(ws, recommendations):
    ws.title = "Recomendacoes"
    headers = ["#", "O Quê (ação)", "Por Quê (evidência)", "Como (formato/hook/CTA)"]
    ws.append(headers)
    for i, r in enumerate(recommendations, 1):
        ws.append([
            i,
            r.get("o_que", ""),
            r.get("por_que", ""),
            r.get("como", ""),
        ])
    style_header(ws)
    autosize(ws, max_width=90)
    # Wrap text para todas as linhas de conteúdo
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", required=True, help="Pasta com dados coletados")
    parser.add_argument("--analysis-json", required=True, help="JSON com análise gerada pelo Claude")
    parser.add_argument("--output", required=True, help="Caminho do .xlsx final")
    args = parser.parse_args()

    base = Path(args.output_dir)

    # Carrega channels
    channels_file = base / "channels.json"
    if not channels_file.exists():
        # Procura num nível acima
        channels_file = base.parent / "channels.json"
    with open(channels_file, "r", encoding="utf-8") as f:
        channels = json.load(f)

    # Carrega videos por canal e top
    videos_by_channel = {}
    all_videos = []
    all_top = []

    for ch in channels:
        slug = ch["slug"]
        ch_dir = base / slug
        if not ch_dir.exists():
            continue

        videos_file = ch_dir / "videos.json"
        if videos_file.exists():
            with open(videos_file, "r", encoding="utf-8") as f:
                videos = json.load(f)
            videos_by_channel[slug] = videos

            top_file = ch_dir / "top_videos.json"
            top_ids = set()
            if top_file.exists():
                with open(top_file, "r", encoding="utf-8") as f:
                    tops = json.load(f)
                top_ids = {v["video_id"] for v in tops}
                all_top.extend(tops)

            for v in videos:
                v["is_top_performer"] = v["video_id"] in top_ids
                all_videos.append(v)

    # Carrega análise gerada pelo Claude
    with open(args.analysis_json, "r", encoding="utf-8") as f:
        analysis = json.load(f)

    # Anexa análise por vídeo aos all_videos
    video_analysis = analysis.get("per_video", {})
    for v in all_videos:
        v["_analysis"] = video_analysis.get(v["video_id"], {})

    # Cria workbook
    wb = Workbook()
    wb.remove(wb.active)

    write_resumo(wb.create_sheet("Resumo"), channels, videos_by_channel, analysis)
    write_videos(wb.create_sheet("Videos"), all_videos)
    write_top_performers(wb.create_sheet("Top_Performers"), all_top, analysis)
    write_comentarios(wb.create_sheet("Comentarios_Categorizados"), analysis.get("categorized_comments", []))
    write_voz_avatar(wb.create_sheet("Voz_Avatar"), analysis.get("voz_avatar", []))
    write_comparativo(wb.create_sheet("Comparativo"), analysis.get("comparativo_matrix", []))
    write_recomendacoes(wb.create_sheet("Recomendacoes"), analysis.get("recomendacoes", []))

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Planilha salva em {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()
